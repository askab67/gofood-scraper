"""
app.py — GoFood Scraper UI
Strategi: _next/data per kecamatan (tidak perlu GeoJSON atau koordinat)
"""

import logging
import os
import threading
import time
import traceback
from pathlib import Path

import pandas as pd
import streamlit as st

# ===========================================================================
# Thread-safe shared state
# ===========================================================================
_shared: dict = {
    "running": False, "completed_points": 0, "total_points": 0,
    "new_restaurants": 0, "total_requests": 0, "error": None,
}
_shared_lock = threading.Lock()

def _set_shared(**kw):
    with _shared_lock: _shared.update(kw)

def _get_shared(k, default=None):
    with _shared_lock: return _shared.get(k, default)

# ===========================================================================
# Logging — persistent file, deduplicated handler
# ===========================================================================
_APP_DIR    = Path(__file__).parent
_LOG_FILE   = str(_APP_DIR / "logs" / "scraper.log")
_CONFIG_DIR = _APP_DIR / "config"
(_APP_DIR / "logs").mkdir(exist_ok=True)
_CONFIG_DIR.mkdir(exist_ok=True)

class _QH(logging.Handler):
    def emit(self, record):
        msg = self.format(record)
        try:
            with open(_LOG_FILE, "a", encoding="utf-8") as f:
                f.write(msg + "\n")
        except Exception:
            pass

_root = logging.getLogger()
if not any(getattr(h, "name", "") == "GFQueueHandler" for h in _root.handlers):
    _h = _QH()
    _h.name = "GFQueueHandler"
    _h.setFormatter(logging.Formatter("%(asctime)s %(levelname)-8s %(message)s", "%H:%M:%S"))
    _root.addHandler(_h)
_root.setLevel(logging.INFO)
_HEADERS_PATH = str(_CONFIG_DIR / "headers.json")

from deduplicate import RestaurantStore
from scraper import GoFoodScraper, BANGKALAN_KECAMATAN

# ===========================================================================
# Daftar area yang sudah diketahui
# Format: {nama_tampil: {"service_area": "...", "localities": [...]}}
# Tambah area baru di sini
# ===========================================================================
KNOWN_AREAS = {
    "Kabupaten Bangkalan (Madura)": {
        "service_area": "madura",
        "localities": [
            "bangkalan", "kamal", "burneh", "socah", "arosbaya",
            "blega", "galis", "geger", "klampis", "kokop",
            "kwanyar", "labang", "modung", "sepulu", "tanah-merah",
            "tanjung-bumi", "tragah",
        ],
    },
    "Kabupaten Sampang (Madura)": {
        "service_area": "madura",
        "localities": [
            "sampang", "camplong", "sreseh", "torjun", "pangarengan",
            "jrengik", "tambelangan", "banyuates", "robatal", "sokobanah",
            "kedungdung", "karang-penang", "omben", "ketapang",
        ],
    },
    "Kabupaten Pamekasan (Madura)": {
        "service_area": "madura",
        "localities": [
            "pamekasan", "pademawu", "galis", "larangan", "pakong",
            "waru", "batumarmar", "pasean", "palengaan", "proppo",
            "kadur", "pegantenan", "tlanakan", "pademawu",
        ],
    },
    "Kabupaten Sumenep (Madura)": {
        "service_area": "madura",
        "localities": [
            "sumenep", "kota-sumenep", "kalianget", "manding",
            "talango", "saronggi", "giligenting", "bluto", "pragaan",
            "lenteng", "ganding", "guluk-guluk", "pasongsongan",
        ],
    },
    "Kota Surabaya": {
        "service_area": "surabaya",
        "localities": [
            "wonokromo", "wonocolo", "wiyung", "genteng", "gubeng",
            "gunung-anyar", "sukolilo", "tambaksari", "simokerto",
            "pabean-cantian", "semampir", "krembangan", "kenjeran",
            "bulak", "mulyorejo", "sawahan", "gayungan", "jambangan",
            "karang-pilang", "dukuh-pakis", "wiyung", "benowo",
            "pakal", "asemrowo", "sukomanunggal", "tandes", "sambikerep",
            "lakarsantri",
        ],
    },
    "⚙️ Area Kustom (masukkan manual)": {
        "service_area": "",
        "localities": [],
    },
}

# ===========================================================================
# Areas manager — baca/tulis config/areas.json
# ===========================================================================
_AREAS_FILE = _CONFIG_DIR / "areas.json"

def _load_custom_areas() -> dict:
    """Baca area kustom dari config/areas.json."""
    import json as _j
    if _AREAS_FILE.exists():
        try:
            return _j.loads(_AREAS_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def _save_custom_areas(areas: dict):
    """Simpan area kustom ke config/areas.json."""
    import json as _j
    _CONFIG_DIR.mkdir(exist_ok=True)
    _AREAS_FILE.write_text(
        _j.dumps(areas, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

def _get_all_areas() -> dict:
    """
    Gabungkan KNOWN_AREAS (bawaan) + custom areas dari areas.json.
    Custom areas ditampilkan di atas, sebelum area bawaan.
    """
    custom = _load_custom_areas()
    combined = {}
    if custom:
        combined.update(custom)
    # Area bawaan — kecuali "Area Kustom" yang dipindah ke bawah
    for k, v in KNOWN_AREAS.items():
        if not k.startswith("⚙️"):
            combined[k] = v
    combined["⚙️ Area Kustom (masukkan manual)"] = KNOWN_AREAS["⚙️ Area Kustom (masukkan manual)"]
    return combined

# ===========================================================================
# Page config & session state
# ===========================================================================
st.set_page_config(
    page_title="GoFood Scraper",
    page_icon="🍴",
    layout="wide",
    initial_sidebar_state="expanded",
)

for _k, _v in {
    "store": None, "scraper_ref": None, "stop_event": None,
    "thread": None, "csv_path": "data/result.csv",
}.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ===========================================================================
# Sidebar
# ===========================================================================
with st.sidebar:
    st.title("⚙️ Konfigurasi")
    st.markdown("---")

    # 1. API Headers
    st.subheader("1. API Headers")
    # ── Fungsi ambil cookie dari Chrome ─────────────────────────────────────
    def _ambil_cookie_chrome() -> str:
        """
        Baca cookie gofood.co.id dari Chrome tanpa perlu admin.
        Coba rookiepy dulu (handles Windows encryption), fallback ke browser_cookie3.
        """
        # ── Coba rookiepy (tidak butuh admin di Windows) ─────────────────
        try:
            import rookiepy
            jar = rookiepy.chrome(["gofood.co.id"])
            cookies = {c["name"]: c["value"] for c in jar}
            if cookies:
                return "; ".join(f"{k}={v}" for k, v in cookies.items())
        except ImportError:
            pass   # rookiepy belum install, coba cara lain
        except Exception:
            pass

        # ── Coba browser_cookie3 ─────────────────────────────────────────
        try:
            import browser_cookie3
            jar = browser_cookie3.chrome(domain_name=".gofood.co.id")
            cookies = {c.name: c.value for c in jar}
            if not cookies:
                jar = browser_cookie3.chrome(domain_name="gofood.co.id")
                cookies = {c.name: c.value for c in jar}
            if cookies:
                return "; ".join(f"{k}={v}" for k, v in cookies.items())
        except ImportError:
            pass
        except Exception as e:
            err = str(e)
            if "admin" in err.lower() or "access" in err.lower():
                return "ERROR:admin"
            return f"ERROR:{e}"

        # ── Cek apakah library sama sekali belum ada ─────────────────────
        try:
            import rookiepy
        except ImportError:
            try:
                import browser_cookie3
            except ImportError:
                return "ERROR:install"

        return "ERROR:empty"

    def _simpan_headers(cookie_str: str):
        """Buat headers.json dari cookie string."""
        import json as _json
        headers = {
            "accept":           "application/json, text/plain, */*",
            "accept-language":  "id-ID,id;q=0.9,en-US;q=0.8",
            "user-agent":       "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
            "referer":          "https://gofood.co.id/",
            "origin":           "https://gofood.co.id",
            "tenant":           "gofood",
            "_activity_source": "gofood",
            "_cookie":          cookie_str,
            "_note":            "Auto-generated. Klik 'Ambil Cookie' lagi jika expired (tiap 1-4 jam).",
        }
        _CONFIG_DIR.mkdir(exist_ok=True)
        (_CONFIG_DIR / "headers.json").write_text(
            _json.dumps(headers, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )

    # ── UI headers ───────────────────────────────────────────────────────────
    st.subheader("1. Cookie GoFood")
    st.caption("Cookie diperlukan agar GoFood mengizinkan akses data.")

    # Tombol utama — ambil otomatis dari Chrome
    if st.button("🍪 Ambil Cookie dari Chrome", use_container_width=True, type="primary"):
        result = _ambil_cookie_chrome()
        if result == "ERROR:install":
            st.error(
                "Library belum terinstall. Jalankan di terminal:\n\n"
                "```\npip install browser-cookie3\n```\n"
                "Lalu restart app dan klik tombol ini lagi."
            )
        elif result == "ERROR:empty":
            st.warning(
                "Cookie GoFood tidak ditemukan di Chrome.\n\n"
                "Pastikan:\n"
                "1. Sudah buka **gofood.co.id** di Chrome\n"
                "2. Browser Chrome **dalam keadaan tertutup** saat klik tombol ini\n"
                "3. Coba tutup Chrome dulu, lalu klik tombol ini lagi"
            )
        elif result == "ERROR:admin":
            st.error(
                "Chrome mengenkripsi cookie dan butuh akses khusus.\n\n"
                "**Solusi:**\n"
                "1. Buka terminal baru **sebagai Administrator**\n"
                "   (klik kanan Command Prompt → 'Run as administrator')\n"
                "2. Jalankan: `pip install rookiepy`\n"
                "3. Restart app dan coba lagi\n\n"
                "Atau gunakan **Upload manual** di bawah."
            )
            st.error(f"Gagal baca cookie: {result[6:]}")
        else:
            _simpan_headers(result)
            n_cookies = result.count(";") + 1
            st.success(f"✓ Cookie berhasil diambil ({n_cookies} cookies) dan disimpan!")

    # Status cookie saat ini
    if Path(_HEADERS_PATH).exists():
        import json as _json2
        try:
            _hdr = _json2.loads(Path(_HEADERS_PATH).read_text(encoding="utf-8"))
            _cookie_str = _hdr.get("_cookie", "")
            _n = _cookie_str.count(";") + 1 if _cookie_str else 0
            if _n > 0:
                st.success(f"✓ Cookie aktif ({_n} cookies tersimpan)")
            else:
                st.warning("⚠ headers.json ada tapi cookie kosong")
        except Exception:
            st.warning("⚠ headers.json tidak bisa dibaca")
    else:
        st.warning("⚠ Cookie belum ada — klik tombol di atas")

    # Fallback: upload manual (untuk pengguna non-Chrome atau kalau tombol gagal)
    with st.expander("⚙️ Upload manual (opsional)"):
        st.caption("Gunakan ini jika tombol otomatis gagal atau pakai Firefox.")
        headers_file = st.file_uploader(
            "Upload headers.json",
            type=["json"],
            help="Lihat tab 'Cara Pakai' untuk instruksi capture manual",
        )
        if headers_file:
            _CONFIG_DIR.mkdir(exist_ok=True)
            (_CONFIG_DIR / "headers.json").write_bytes(headers_file.read())
            st.success("headers.json tersimpan ✓")

    st.markdown("---")

    # 2. GeoJSON (untuk fallback kecamatan kecil)
    st.subheader("2. GeoJSON Batas Kecamatan")
    st.caption(
        "Diperlukan untuk kecamatan kecil yang tidak punya halaman GoFood sendiri. "
        "Upload file batas kecamatan Bangkalan."
    )
    geojson_file = st.file_uploader(
        "Upload GeoJSON batas kecamatan",
        type=["geojson", "json"],
    )

    kecamatan_polygons = {}
    if geojson_file:
        try:
            import json as _json
            from polygon_filter import extract_kecamatan_polygons
            gj_dict = _json.loads(geojson_file.read())
            kecamatan_polygons = extract_kecamatan_polygons(gj_dict)
            st.success(
                f"✓ GeoJSON dimuat — "
                f"{len(kecamatan_polygons)} polygon kecamatan tersedia"
            )
            with st.expander("Lihat kecamatan dalam GeoJSON"):
                for k in sorted(kecamatan_polygons.keys()):
                    st.write(f"• {k}")
        except Exception as e:
            st.error(f"Gagal baca GeoJSON: {e}")
    else:
        st.info(
            "Tanpa GeoJSON, kecamatan kecil (burneh, socah, dll) akan dilewati. "
            "Upload untuk scrape semua kecamatan."
        )

    st.markdown("---")
    st.subheader("3. Pilih Daerah")
    _all_areas = _get_all_areas()
    area_name  = st.selectbox("Daerah", list(_all_areas.keys()))
    area_cfg   = _all_areas[area_name]

    if area_name.startswith("⚙️"):
        # Custom area
        custom_service = st.text_input(
            "serviceArea",
            placeholder="contoh: madura, surabaya, jakarta",
            help="Lihat URL GoFood: gofood.co.id/id/{serviceArea}/...",
        )
        custom_locs = st.text_area(
            "Daftar locality (satu per baris)",
            placeholder="bangkalan\nkamal\nsocah",
            help="Nama locality sesuai URL GoFood",
        )
        service_area = custom_service.strip()
        localities   = [l.strip() for l in custom_locs.splitlines() if l.strip()]
    else:
        service_area = area_cfg["service_area"]
        localities   = area_cfg["localities"]
        st.info(
            f"serviceArea: **{service_area}** | "
            f"**{len(localities)}** kecamatan/kelurahan"
        )

        # Opsi pilih kecamatan tertentu saja
        if st.toggle("Pilih kecamatan tertentu saja"):
            localities = st.multiselect(
                "Kecamatan yang di-scrape",
                options=localities,
                default=localities,
            )

    st.markdown("---")

    # 3. Scraper settings
    st.subheader("4. Pengaturan Scraper")
    delay_min   = st.slider("Min delay (s)", 0.5, 3.0, 1.0, 0.1)
    delay_max   = st.slider("Max delay (s)", 1.0, 5.0, 2.5, 0.1)
    max_threads = st.slider("Threads", 1, 4, 2, 1,
                            help="2 threads sudah cukup. Lebih banyak = risiko diblokir")
    max_pages   = st.slider("Max halaman per kategori", 1, 20, 5, 1)

    st.markdown("---")

    # 4. Output
    st.subheader("5. Output")
    out_dir         = st.text_input("Folder output", value="data")
    resume_scraping = st.toggle("Lanjutkan run sebelumnya", value=True)

    with st.expander("Proxy (opsional)"):
        proxy_http  = st.text_input("HTTP proxy",  placeholder="http://host:port")
        proxy_https = st.text_input("HTTPS proxy", placeholder="http://host:port")


# ===========================================================================
# Main area
# ===========================================================================
st.title("🍴 GoFood Restaurant Scraper")
st.caption("Scrape semua restoran GoFood per kecamatan via _next/data endpoint")

tabs = st.tabs(["🚀 Scraper", "📋 Hasil", "🔀 Merge Runs", "🗺 Peta Alamat", "🌍 Kelola Daerah", "ℹ️ Cara Pakai"])


# ============================================================
# TAB 1 – Scraper
# ============================================================
with tabs[0]:
    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.subheader("Target Scraping")
        if localities:
            st.success(
                f"**{len(localities)}** kecamatan/kelurahan dari "
                f"**{area_name}** siap di-scrape"
            )
            with st.expander(f"Lihat daftar ({len(localities)} kecamatan)"):
                cols = st.columns(3)
                for i, loc in enumerate(localities):
                    cols[i % 3].write(f"• {loc}")
        else:
            st.warning("Belum ada kecamatan dipilih. Pilih daerah di sidebar.")

    with col_right:
        st.subheader("Status")
        running   = _get_shared("running")
        total_pts = _get_shared("total_points", 0)
        done_pts  = _get_shared("completed_points", 0)
        n_reqs    = _get_shared("total_requests", 0)
        err       = _get_shared("error")

        thread = st.session_state.get("thread")
        if running and thread and not thread.is_alive():
            _set_shared(running=False)
            running = False

        if running:
            st.warning("⏳ Scraping berjalan…")
        elif err:
            st.error("💥 Error — lihat detail di bawah")
        else:
            st.success("⬤ Siap")

        m1, m2, m3 = st.columns(3)
        store = st.session_state.get("store")
        m1.metric("Restoran",        f"{len(store):,}" if store else "0")
        m2.metric("Kecamatan Selesai", f"{done_pts} / {total_pts}")
        m3.metric("API Requests",      f"{n_reqs:,}")

        if err:
            with st.expander("🔴 Detail error"):
                st.code(err, language=None)

    st.divider()

    # Tombol kontrol
    bc1, bc2, bc3 = st.columns(3)
    can_start = bool(localities) and not running
    start_btn = bc1.button("▶ Mulai Scraping", type="primary", disabled=not can_start)
    stop_btn  = bc2.button("⏹ Stop",           disabled=not running)
    clear_btn = bc3.button("🗑 Hapus Log")

    pct = (done_pts / max(total_pts, 1)) if total_pts > 0 else 0.0
    st.progress(min(pct, 1.0))

    # Log
    st.subheader("📜 Live Log")
    log_display = st.empty()
    if Path(_LOG_FILE).exists():
        try:
            lines = Path(_LOG_FILE).read_text(encoding="utf-8").splitlines()
            log_display.code("\n".join(lines[-80:]), language=None)
        except Exception:
            pass
    else:
        log_display.info("Log akan muncul di sini setelah scraping dimulai.")

    if running:
        st.button("🔄 Refresh log", key="refresh")

    # ── Handler ─────────────────────────────────────────────────────────
    if stop_btn:
        se = st.session_state.get("stop_event")
        if se: se.set()
        _set_shared(running=False)
        st.warning("⏹ Stop dikirim.")

    if clear_btn:
        Path(_LOG_FILE).write_text("", encoding="utf-8")
        st.success("Log dihapus.")

    if start_btn and localities:
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        csv_path  = os.path.join(out_dir, "result.csv")
        json_path = os.path.join(out_dir, "result.json")
        st.session_state["csv_path"] = csv_path

        store = (RestaurantStore.load_or_create(csv_path)
                 if resume_scraping else RestaurantStore())
        st.session_state["store"] = store

        proxies = {}
        if proxy_http:  proxies["http"]  = proxy_http
        if proxy_https: proxies["https"] = proxy_https

        stop_event = threading.Event()
        st.session_state["stop_event"] = stop_event

        _set_shared(
            running=True, completed_points=0,
            total_points=len(localities),
            new_restaurants=0, total_requests=0, error=None,
        )
        Path(_LOG_FILE).write_text("", encoding="utf-8")

        def _on_progress(current, total, new_count):
            _set_shared(completed_points=current, total_points=total)

        scraper = GoFoodScraper(
            store               = store,
            headers_path        = _HEADERS_PATH,
            delay_min           = delay_min,
            delay_max           = delay_max,
            max_threads         = max_threads,
            proxies             = proxies if proxies else None,
            on_progress         = _on_progress,
            stop_event          = stop_event,
            max_pages           = max_pages,
            kecamatan_list      = localities,
            service_area        = service_area,
            kecamatan_polygons  = kecamatan_polygons,
        )
        st.session_state["scraper_ref"] = scraper

        def _run():
            try:
                logging.info(
                    "=== Mulai: %d kecamatan | serviceArea=%s ===",
                    len(localities), service_area,
                )
                scraper.run(save_csv=csv_path, save_json=json_path)
                logging.info("=== Selesai ===")
            except Exception:
                tb = traceback.format_exc()
                logging.error("Crash:\n%s", tb)
                _set_shared(error=tb)
            finally:
                _set_shared(
                    running=False,
                    total_requests=scraper.total_requests,
                    new_restaurants=scraper.new_restaurants,
                )

        t = threading.Thread(target=_run, daemon=True, name="scraper")
        st.session_state["thread"] = t
        t.start()
        st.rerun()

    # ── Auto-refresh while running ───────────────────────────────────────────
    # Baca stats langsung dari scraper object — lebih reliable dari callback
    ref = st.session_state.get("scraper_ref")
    if ref and _get_shared("running"):
        with ref._lock:
            _set_shared(
                total_requests   = ref.total_requests,
                new_restaurants  = ref.new_restaurants,
                completed_points = ref._completed_kecamatan,
                total_points     = len(ref.kecamatan_list),
            )
    if _get_shared("running"):
        time.sleep(3)
        st.rerun()


# ============================================================
# TAB 2 – Hasil
# ============================================================
with tabs[1]:
    st.subheader("📋 Hasil Scraping")
    result_csv = Path(st.session_state.get("csv_path", "data/result.csv"))

    if result_csv.exists():
        df = pd.read_csv(result_csv)
        st.success(f"**{len(df):,}** restoran di `{result_csv}`")

        # Filter
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            is_open_opts = ["Semua"] + sorted(df["is_open_status"].dropna().unique().tolist())
            open_filter  = st.selectbox("Status", is_open_opts)
        with fc2:
            rating_min = st.slider("Min rating", 0.0, 5.0, 0.0, 0.1)
        with fc3:
            price_opts  = ["Semua"] + sorted(df["price_range"].dropna().unique().tolist())
            price_filter = st.selectbox("Harga", price_opts)

        filtered = df.copy()
        if open_filter != "Semua":
            filtered = filtered[filtered["is_open_status"] == open_filter]
        if rating_min > 0:
            filtered = filtered[filtered["rating"] >= rating_min]
        if price_filter != "Semua":
            filtered = filtered[filtered["price_range"] == price_filter]

        st.dataframe(filtered, use_container_width=True, height=420)
        st.caption(f"Menampilkan {len(filtered):,} dari {len(df):,} restoran")

        dc1, dc2 = st.columns(2)
        dc1.download_button(
            "⬇ Download CSV",
            data=filtered.to_csv(index=False).encode("utf-8"),
            file_name="gofood_result.csv",
            mime="text/csv",
        )

        # Excel download via openpyxl jika tersedia
        try:
            import openpyxl
            from io import BytesIO
            from openpyxl.styles import Font, PatternFill, Alignment

            def _to_excel(dataframe):
                buf = BytesIO()
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "GoFood"
                hfill = PatternFill("solid", fgColor="1B5E20")
                hfont = Font(bold=True, color="FFFFFF")
                col_w = [38,45,25,8,12,40,12,12,55,10,12,10,65]
                for i, row in enumerate([filtered.columns.tolist()] + filtered.values.tolist(), 1):
                    for j, val in enumerate(row, 1):
                        c = ws.cell(row=i, column=j, value=val)
                        if i == 1:
                            c.font = hfont
                            c.fill = hfill
                            c.alignment = Alignment(horizontal="center")
                        if j <= len(col_w):
                            ws.column_dimensions[
                                openpyxl.utils.get_column_letter(j)
                            ].width = col_w[j-1]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                wb.save(buf)
                return buf.getvalue()

            dc2.download_button(
                "⬇ Download Excel (.xlsx)",
                data=_to_excel(filtered),
                file_name="gofood_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ImportError:
            dc2.caption("Install `openpyxl` untuk download Excel")

        if st.toggle("Tampilkan di peta"):
            map_df = filtered.dropna(subset=["latitude","longitude"]).rename(
                columns={"latitude":"lat","longitude":"lon"}
            )
            if not map_df.empty:
                st.map(map_df[["lat","lon"]])
    else:
        st.info("Belum ada hasil. Jalankan scraper terlebih dahulu.")


# ============================================================
# TAB 3 – Merge
# ============================================================
with tabs[2]:
    st.subheader("🔀 Gabungkan Hasil Beberapa Run")
    st.caption("Upload 2+ file CSV dari run yang berbeda. Duplikat akan otomatis dihapus.")

    merge_files = st.file_uploader(
        "Upload file CSV", type=["csv"], accept_multiple_files=True
    )
    merge_name  = st.text_input("Nama file output", value="result_merged.csv")

    if st.button("🔀 Gabungkan", disabled=len(merge_files) < 2):
        dfs = []
        for f in merge_files:
            df_tmp = pd.read_csv(f)
            dfs.append(df_tmp)
            st.write(f"• `{f.name}` — {len(df_tmp):,} baris")

        merged = (pd.concat(dfs, ignore_index=True)
                  .drop_duplicates(subset=["restaurant_id"])
                  .reset_index(drop=True))
        st.success(f"Hasil gabungan: **{len(merged):,}** restoran unik")
        st.dataframe(merged, use_container_width=True)
        st.download_button(
            "⬇ Download CSV gabungan",
            data=merged.to_csv(index=False).encode("utf-8"),
            file_name=merge_name,
            mime="text/csv",
        )


# ============================================================
# TAB 4 – Reverse Geocode (isi kolom address)
# ============================================================
with tabs[3]:
    st.subheader("🗺 Isi Kolom Address dari Koordinat")
    st.markdown(
        "Karena GoFood tidak menyediakan alamat teks, kita bisa mengisi kolom "
        "`address` secara otomatis menggunakan **koordinat latitude/longitude** "
        "yang sudah ada melalui layanan **Nominatim (OpenStreetMap)** — gratis, "
        "tanpa API key."
    )

    rg_csv = st.file_uploader("Upload result.csv", type=["csv"], key="rg_upload")

    if rg_csv:
        df_rg = pd.read_csv(rg_csv)
        n_total   = len(df_rg)
        n_missing = df_rg["address"].isna().sum() + (df_rg["address"] == "").sum()
        st.info(f"Total: **{n_total}** restoran | Address kosong: **{n_missing}**")

        delay_rg = st.slider("Delay antar request (s)", 1.0, 3.0, 1.1, 0.1,
                             help="Nominatim mensyaratkan min 1 request/detik")
        limit_rg = st.number_input("Maks restoran yang diproses (0=semua)", 0, n_total, 0)

        if st.button("🗺 Mulai reverse geocode"):
            import requests as _req

            needs = df_rg[
                (df_rg["address"].isna() | (df_rg["address"] == "")) &
                df_rg["latitude"].notna() & df_rg["longitude"].notna()
            ].copy()

            if limit_rg > 0:
                needs = needs.head(int(limit_rg))

            prog  = st.progress(0.0)
            info  = st.empty()
            total = len(needs)
            done  = 0
            errors = 0

            NOM_URL = "https://nominatim.openstreetmap.org/reverse"
            NOM_HDR = {"User-Agent": "GoFoodScraper/1.0 (research)"}

            for idx, row in needs.iterrows():
                try:
                    r = _req.get(NOM_URL, params={
                        "lat": row["latitude"], "lon": row["longitude"],
                        "format": "json", "addressdetails": 1,
                        "zoom": 18,
                    }, headers=NOM_HDR, timeout=10)
                    r.raise_for_status()
                    result = r.json()
                    display = result.get("display_name", "")
                    df_rg.at[idx, "address"] = display
                except Exception:
                    errors += 1

                done += 1
                prog.progress(done / max(total, 1))
                info.write(
                    f"Diproses: {done}/{total} | "
                    f"Error: {errors} | "
                    f"Terakhir: {df_rg.at[idx, 'restaurant_name'][:35]}"
                )
                time.sleep(delay_rg)

            prog.progress(1.0)
            st.success(f"Selesai! {done - errors} alamat berhasil diisi.")
            st.dataframe(df_rg[["restaurant_name","address","latitude","longitude"]].head(20))
            st.download_button(
                "⬇ Download CSV dengan alamat",
                data=df_rg.to_csv(index=False).encode("utf-8"),
                file_name="result_with_address.csv",
                mime="text/csv",
            )


# ============================================================
# TAB 5 – Kelola Daerah
# ============================================================
with tabs[4]:
    st.subheader("🌍 Kelola Daerah")
    st.caption(
        "Tambah atau hapus daerah scraping tanpa perlu edit kode. "
        "Data disimpan di `config/areas.json`."
    )

    custom_areas = _load_custom_areas()

    # ── Form tambah daerah baru ──────────────────────────────────────────
    st.markdown("### ➕ Tambah Daerah Baru")

    with st.form("form_tambah_daerah", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            nama_baru = st.text_input(
                "Nama daerah",
                placeholder="contoh: Kabupaten Gresik",
                help="Nama yang akan muncul di dropdown pilih daerah",
            )
            service_area_baru = st.text_input(
                "serviceArea",
                placeholder="contoh: surabaya",
                help=(
                    "Lihat URL GoFood saat pilih lokasi: "
                    "gofood.co.id/id/**surabaya**/gresik-restaurants"
                ),
            )
        with col2:
            localities_baru = st.text_area(
                "Daftar kecamatan/kelurahan (satu per baris)",
                placeholder="gresik\nkebomas\nmanyar\nbungah\nsidayu",
                height=150,
                help=(
                    "Nama locality sesuai URL GoFood: "
                    "gofood.co.id/id/surabaya/**gresik**-restaurants"
                ),
            )

        # Cara cari serviceArea dan locality
        with st.expander("💡 Cara cari serviceArea dan locality"):
            st.markdown("""
1. Buka **gofood.co.id** di browser
2. Klik ikon lokasi → pilih kota/kabupaten yang diinginkan
3. Lihat URL yang terbentuk:
   ```
   gofood.co.id/id/{serviceArea}/{locality}-restaurants
   ```
4. Contoh untuk Gresik:
   ```
   gofood.co.id/id/surabaya/gresik-restaurants
                    ↑           ↑
              serviceArea     locality
   ```
5. Untuk kecamatan lain, ubah nama locality-nya saja
""")

        submitted = st.form_submit_button("➕ Tambah Daerah", type="primary", use_container_width=True)

        if submitted:
            if not nama_baru.strip():
                st.error("Nama daerah tidak boleh kosong.")
            elif not service_area_baru.strip():
                st.error("serviceArea tidak boleh kosong.")
            elif not localities_baru.strip():
                st.error("Minimal isi 1 kecamatan/kelurahan.")
            elif nama_baru.strip() in KNOWN_AREAS:
                st.error(f"Nama '{nama_baru}' sudah ada di daftar bawaan.")
            elif nama_baru.strip() in custom_areas:
                st.error(f"Nama '{nama_baru}' sudah ada. Hapus dulu jika ingin mengganti.")
            else:
                locs = [
                    l.strip().lower().replace(" ", "-")
                    for l in localities_baru.splitlines()
                    if l.strip()
                ]
                custom_areas[nama_baru.strip()] = {
                    "service_area": service_area_baru.strip().lower(),
                    "localities":   locs,
                }
                _save_custom_areas(custom_areas)
                st.success(
                    f"✓ **{nama_baru.strip()}** berhasil ditambahkan "
                    f"({len(locs)} kecamatan). "
                    f"Sekarang tersedia di dropdown 'Pilih Daerah'."
                )
                st.rerun()

    st.divider()

    # ── Daftar daerah yang sudah ada ─────────────────────────────────────
    st.markdown("### 📋 Daftar Semua Daerah")

    # Daerah bawaan (read-only)
    st.markdown("**Daerah bawaan** (tidak bisa dihapus):")
    builtin_names = [k for k in KNOWN_AREAS if not k.startswith("⚙️")]
    cols = st.columns(2)
    for i, name in enumerate(builtin_names):
        locs = KNOWN_AREAS[name]["localities"]
        cols[i % 2].info(f"**{name}**  \n{len(locs)} kecamatan")

    # Daerah kustom (bisa dihapus)
    if custom_areas:
        st.markdown("**Daerah kustom** (bisa dihapus):")
        for nama, cfg in list(custom_areas.items()):
            c1, c2, c3 = st.columns([3, 1, 1])
            c1.write(f"**{nama}**")
            c2.write(f"{len(cfg['localities'])} kecamatan | `{cfg['service_area']}`")
            if c3.button("🗑 Hapus", key=f"del_{nama}"):
                del custom_areas[nama]
                _save_custom_areas(custom_areas)
                st.success(f"'{nama}' dihapus.")
                st.rerun()

        # Detail kecamatan per daerah kustom
        with st.expander("Lihat detail kecamatan per daerah kustom"):
            for nama, cfg in custom_areas.items():
                st.markdown(f"**{nama}** (`{cfg['service_area']}`):")
                st.write(", ".join(cfg["localities"]))
    else:
        st.info("Belum ada daerah kustom. Tambahkan menggunakan form di atas.")


# ============================================================
# TAB 6 – Cara Pakai
# ============================================================
with tabs[5]:
    st.subheader("ℹ️ Cara Pakai")
    st.markdown("""
### Cara kerja scraper

GoFood menyimpan daftar restoran per kecamatan di endpoint:
```
GET gofood.co.id/_next/data/{buildId}/id/{serviceArea}/{locality}-restaurants/{kategori}.json
```

Scraper:
1. Buka halaman utama kecamatan → dapat daftar kategori (Aneka Nasi, Ayam, Seafood, dll)
2. Buka setiap halaman kategori → dapat daftar restoran
3. Deduplikasi by `uid`, simpan ke CSV

---

### Cara refresh cookie (jika dapat error 401/403)

1. Buka **gofood.co.id** di Chrome
2. F12 → Network → Fetch/XHR
3. Klik salah satu restoran atau search
4. Klik kanan request apa saja → **Copy → Copy as cURL (bash)**
5. Ekstrak nilai dari `-b '...'` (bagian cookie)
6. Buat file `config/headers.json`:
```json
{
  "accept": "application/json, text/plain, */*",
  "accept-language": "id-ID,id;q=0.9",
  "user-agent": "Mozilla/5.0 ...",
  "_cookie": "OptanonConsent=...; csrfSecret=...; XSRF-TOKEN=...; w_tsfp=..."
}
```
7. Upload via sidebar atau taruh langsung di folder `config/`

---

### Cara tambah daerah baru

Edit file `app.py`, tambahkan ke dict `KNOWN_AREAS`:
```python
"Kabupaten Gresik": {
    "service_area": "surabaya",
    "localities": ["gresik", "kebomas", "manyar", "bungah", ...],
},
```

Untuk cari nama `serviceArea` dan `locality`:
1. Buka gofood.co.id, pilih lokasi yang diinginkan
2. Lihat URL yang terbentuk:
   `gofood.co.id/id/{serviceArea}/{locality}-restaurants`

---

### Kenapa address kosong?

GoFood tidak menyediakan alamat teks di endpoint ini (`core.address = null`).
Gunakan tab **🗺 Peta Alamat** untuk mengisi otomatis dari koordinat via Nominatim.

---

### Tips waktu scraping

- **Siang hari**: restoran yang buka siang muncul
- **Malam / jam berbuka puasa**: restoran yang buka malam muncul
- **Jalankan 2x** (siang + malam) lalu **Merge** untuk hasil paling lengkap
""")

    st.divider()
    st.subheader("📄 File Log Lengkap")
    if Path(_LOG_FILE).exists():
        log_text = Path(_LOG_FILE).read_text(encoding="utf-8")
        st.text_area("logs/scraper.log", log_text, height=350)
        st.download_button(
            "⬇ Download log",
            data=log_text.encode("utf-8"),
            file_name="scraper.log",
            mime="text/plain",
        )
    else:
        st.info("Log belum ada.")