import csv
import sys
from pathlib import Path

# ── Coba buat Excel ──────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    CSV_PATH  = Path("data/result.csv")
    XLSX_PATH = Path("data/result.xlsx")

    if not CSV_PATH.exists():
        print(f"File tidak ditemukan: {CSV_PATH}")
        sys.exit(1)

    with open(CSV_PATH, "r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))

    if not rows:
        print("CSV kosong!")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GoFood Bangkalan"

    # Header style
    header_fill = PatternFill("solid", fgColor="2E7D32")  # hijau gelap
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Lebar kolom yang pas
    col_widths = {
        "A": 38,   # restaurant_id
        "B": 45,   # restaurant_name
        "C": 25,   # category
        "D": 8,    # rating
        "E": 12,   # review_count
        "F": 40,   # address
        "G": 12,   # latitude
        "H": 12,   # longitude
        "I": 50,   # opening_hours
        "J": 12,   # menu_count
        "K": 12,   # price_range
        "L": 10,   # is_open_status
        "M": 60,   # resto_url
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Baris zebra (selang-seling warna)
    light_green = PatternFill("solid", fgColor="F1F8E9")
    for r_idx in range(2, ws.max_row + 1):
        if r_idx % 2 == 0:
            for c_idx in range(1, ws.max_column + 1):
                ws.cell(row=r_idx, column=c_idx).fill = light_green

    wb.save(XLSX_PATH)
    print(f"✓ Berhasil! File disimpan: {XLSX_PATH}")
    print(f"  Total baris: {ws.max_row - 1} restoran")
    print(f"  Buka dengan Excel atau Google Sheets")

except ImportError:
    print("openpyxl belum terinstall.")
    print("Jalankan: pip install openpyxl")
    print()
    print("Atau buka CSV dengan cara yang benar di Excel:")
    print()
    print("CARA BUKA CSV DI EXCEL:")
    print("1. Buka Excel → tab Data")
    print("2. Klik 'From Text/CSV'")
    print("3. Pilih file result.csv")
    print("4. Delimiter: pilih 'Comma'")
    print("5. Klik Load")
    print()
    print("ATAU di Google Sheets:")
    print("1. File → Import → Upload result.csv")
    print("2. Import location: Replace spreadsheet")
    print("3. Separator: Comma")
    print("4. Klik Import")
